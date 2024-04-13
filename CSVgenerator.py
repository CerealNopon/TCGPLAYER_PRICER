from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from openpyxl.drawing.image import Image as XLImage
import os
import time

# Get list of urls from text file
def getUrls():
    # Open the text file
    with open('INPUT_LINKS_HERE.txt', 'r') as file:
        lines = file.readlines()
    urls = []

    # Iterate through each line
    for line in lines:
        # Strip any leading/trailing whitespace and append the desired text
        modified_url = line.strip() + "&Condition=Lightly+Played"
        # modified_url = line.strip()

        # Append the modified URL to the list
        urls.append(modified_url)

    return urls

# Get number of files
def count_files(folderPath):
    # Ensure the folder exists
    if not os.path.exists(folderPath):
        print("Folder does not exist.")
        return
    
    # Get list of files in the folder
    files = os.listdir(folderPath)
    
    # Count the number of files
    num_files = len(files)
    
    return num_files

# Clear files in folder
def delete_files_in_folder(folder_path):
    # Ensure the folder exists
    if not os.path.exists(folder_path):
        # print("Folder does not exist.")
        return
    
    # Get list of files in the folder
    files = os.listdir(folder_path)
    
    # Iterate over each file and delete it
    for file in files:
        file_path = os.path.join(folder_path, file)
        try:
            if os.path.isfile(file_path):
                os.remove(file_path)
                # print(f"Deleted file: {file_path}")
        except Exception as e:
            # print(f"Error deleting file: {file_path}", e)
            print()

# Define a function to wait for the presence of a class
def wait_for_class(driver, class_name, timeout=10):
    return WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((By.CLASS_NAME, class_name))
    )

# Get name of listing
def getNamePriceAndImage(driver, url):
    driver.get(url)
    try:
        name = wait_for_class(driver, "product-details__name")
        time.sleep(2)
        price = wait_for_class(driver, "listing-item__price")
        image = wait_for_class(driver, "lazy-image__wrapper")
        folderPath = r'C:\Users\rober\TCGplayer_Pricer\images'
        imageFilename = "new" + str(count_files(folderPath)) + ".png"
        image.screenshot(os.path.join(folderPath, imageFilename))
        return name.text, price.text, str(os.path.join(folderPath, imageFilename))
    except Exception as e:
        print("class not found")
        return

def main():

    # Create Driver & Base Link
    chromeOptions = Options()
    chromeOptions.headless = False
    driver = webdriver.Chrome(options=chromeOptions)

    urls = getUrls() # Initialize URLs
    delete_files_in_folder("images") # Clear images folder

    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Listings"

    # Write headers
    headers = ['Name', 'Condition', 'Price', '70% Price', 'Offer', 'Image', 'Link']
    sheet.append(headers)

    for url in urls:
        name, price, imagePath = getNamePriceAndImage(driver, url)
        priceFloat = float(price.replace('$', ''))

        img = XLImage(imagePath)
        img.width = img.width * 0.25  # Resize width to 25% of its original size
        img.height = img.height * 0.25  # Resize height to 25% of its original size
        img.anchor = f'F{sheet.max_row + 1}'  # Place image in the E column, next to the other data
        sheet.add_image(img)

        reducedPrice = round(priceFloat*0.65, 2)

        if reducedPrice < 10:
            offer = int(reducedPrice)
            if offer == 0:
                offer = reducedPrice
        else:
            offer = reducedPrice
        condition = "LP"

        # Name, Condition, Price, 70% Price, Offer, Image, Link
        sheet.append([name, condition, priceFloat, reducedPrice, offer, "", url])

    sheet.cell(row=sheet.max_row + 1, column=2, value="Total:")

    # Save the workbook
    workbook.save(filename='listings.xlsx')

    driver.quit()

if __name__ == "__main__":
    main()
